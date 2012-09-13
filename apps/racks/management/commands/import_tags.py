from django.core.management.base import NoArgsCommand
from openpyxl import load_workbook

from apps.tagging.models import Tag

def import_tags():    
    wb = load_workbook(filename = r'Tags.xlsx')
    
    ws = wb.get_sheet_by_name(name = 'Sheet1')
    for rownum in range(ws.get_highest_row()):
        if rownum >= 3:
            tag_name = ws.cell(row=rownum, column=0).value
            if Tag.objects.filter(name=tag_name).count() > 0:
                pass
            else:
                Tag.objects.create(name=tag_name)
            

class Command(NoArgsCommand):
    help = ""

    def handle_noargs(self, **options):
        import_tags()