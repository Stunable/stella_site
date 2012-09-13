from django.core.management.base import NoArgsCommand
from openpyxl import load_workbook

from racks.models import PriceCategory, Brand

def import_jean_brands():    
    wb = load_workbook(filename = r'Jean_Logic_2.23.xlsx')
    
    ws = wb.get_sheet_by_name(name = 'Sheet1')
    for rownum in range(ws.get_highest_row()):
        if rownum >= 1:
            brand_name = ws.cell(row=rownum, column=0).value
            brand_price_category = ws.cell(row=rownum, column=1).value
            if brand_price_category == 1:
                price_category = PriceCategory.objects.get(name="Category A")
            elif brand_price_category == 2:
                price_category = PriceCategory.objects.get(name="Category B")
            elif brand_price_category == 3:
                price_category = PriceCategory.objects.get(name="Category C")
            elif brand_price_category == 4:
                price_category = PriceCategory.objects.get(name="Category D")
            elif brand_price_category == 5:
                price_category = PriceCategory.objects.get(name="Category E")
            else:
                price_category = None
            
            Brand.objects.create(name=brand_name, website="", category=price_category)
            

class Command(NoArgsCommand):
    help = ""

    def handle_noargs(self, **options):
        import_jean_brands()