from django.core.management.base import NoArgsCommand
from openpyxl import load_workbook
from django.core.files.temp import NamedTemporaryFile
from django.core.files import File

from racks.models import Item, Category, Size, Color, ItemType
from apps.retailers.models import StylistItem, RetailerProfile

def import_products():    
    wb = load_workbook(filename = r'Product Upload Sheet 2.xlsx')
    brand = "Andrea Bocchio"
    name = None
    description = None
    price = None
    product_category = None
    product_image_id = None
    price = None
    product_color = None
    product_size = None
    
    ws = wb.get_sheet_by_name(name = 'Master')
    for rownum in xrange(1, ws.get_highest_row()):
        item = Item()
#        import pdb; pdb.set_trace()
        
#        if not ws.cell(row=rownum, column=1).value or not ws.cell(row=rownum, column=2).value:
#            continue
        
        brand = ws.cell(row=rownum, column=0).value or brand
        
        name = ws.cell(row=rownum, column=1).value or name
        
        description = ws.cell(row=rownum, column=3).value or name
        
        price = ws.cell(row=rownum, column=8).value or price
        
        product_category = ws.cell(row=rownum, column=4).value or product_category
        
        product_image_id = ws.cell(row=rownum, column=2).value or product_image_id
        
        if type(product_image_id) == unicode:
            product_image_id = product_image_id.strip("'")
        else:
            product_image = '%04d.jpg' % product_image_id

        
        print rownum, product_image
        image_file = open('Images/'+ product_image, 'r')
        img_temp = NamedTemporaryFile(delete=True)
        img_temp.write(image_file.read())
        img_temp.flush()
        
        item.name = name
        
        item.description = description
        
        item.brand = brand
    
        
        print price
        
        item.price = price or 0.0
        
        try:
            category = Category.objects.get(name=product_category)
        except Category.DoesNotExist, e:
            print product_category, 'does not exist'
            Category(name=product_category).save()
            raise(e)
        
        
        item.category = category
        item.save()
        
        print 'brand', brand
        retailer_profile = RetailerProfile.objects.filter(name=brand)[0]
        StylistItem(stylist=retailer_profile.user, item=item).save()
        
        item.image.save(product_image, File(img_temp))
        
        product_color = ws.cell(row=rownum, column=6).value or product_color
        
        if Color.objects.filter(name=product_color).count() > 0:
            pass
        else:
            Color.objects.create(name=product_color)
            
        print product_color
        
        color = Color.objects.get(name=product_color)
        
        product_size = ws.cell(row=rownum, column=7).value or product_size
        
        for size in product_size.split(','):
            size = size.strip()
        
            if size == "One size" or size == 'One Size':
                size = Size.objects.get(size="One Size Fits All")
            else:
                try:
                    size = Size.objects.get(size=size)
                except:
                    size = Size(size=size)
                    size.save()
                
            item_type = ItemType()
            item_type.item = item
            item_type.size = size
            item_type.color = color
            item_type.save()
            
            print (size, color)
                
            
            

class Command(NoArgsCommand):
    help = ""

    def handle_noargs(self, **options):
        import_products()